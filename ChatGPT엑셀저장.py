# ChatGPT엑셀저장.py
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import random
from datetime import datetime, timedelta

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Define column headers
headers = ["Product Name", "Price", "Quantity", "Sale Date"]

# Add headers to the first row and apply bold font
for col_num, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)

# Generate and add sales data
for row_num in range(2, 102):  # Generate 100 pieces of data
    product_name = f"Product {row_num - 1}"
    price = round(random.uniform(100, 1000), 2)
    quantity = random.randint(1, 10)
    sale_date = datetime(2023, 1, 1) + timedelta(days=random.randint(1, 365))

    data = [product_name, price, quantity, sale_date]

    for col_num, value in enumerate(data, start=1):
        cell = sheet.cell(row=row_num, column=col_num)
        cell.value = value

# Save the workbook to the specified path
file_path = "c:\\work\\sales.xlsx"
workbook.save(file_path)

print(f"Sales data saved to {file_path}")
